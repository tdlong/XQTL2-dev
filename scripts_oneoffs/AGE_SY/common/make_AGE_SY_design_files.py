#!/usr/bin/env python3
"""
make_AGE_SY_design_files.py — Generate AGE_SY pipeline design files from summary_info_v1.xlsx

Reads helpfiles/AGE_SY/summary_info_v1.xlsx (two sheets):

  Control_Flies:
    Replicate           — Rep1, Rep2, ... (skip RepA* rows — different base population)
    Sex                 — F or M
    Num_Random_Controls — fly count (used as Num in design file)

  Aged_Flies:
    Replicate           — Rep1, Rep2, ... (skip RepA* rows)
    Sex                 — F or M
    Diet                — SY10 or SY20
    Num_Aged            — fly count (used as Num in design file)
    Percent_Aged        — divide by 100 → Proportion in design file

BAM name convention: {longTRT}_R{n}_{sex}
  e.g. AgeSY10_R8_F, AgeSY20_R9_M, Con_R10_F

Output files (in helpfiles/AGE_SY/):
  AGE_SY10_F.test.txt  — AgeSY10 vs Con, females, R1-R12
  AGE_SY10_M.test.txt  — AgeSY10 vs Con, males,   R1-R12
  AGE_SY20_F.test.txt  — AgeSY20 vs Con, females, R1-R12
  AGE_SY20_M.test.txt  — AgeSY20 vs Con, males,   R1-R12

IMPORTANT: filesize (column 2) is written as 0 — a placeholder.
           After the BAMs are merged, update with actual sizes:
             stat -c "%s" data/bam/AGE_SY/<sample>.bam   (Linux)
           Then replace the 0s in each design file (or rerun this script
           with the --filesizes option once that is implemented).

Run from: XQTL2-dev root
  python3 scripts_oneoffs/AGE_SY/common/make_AGE_SY_design_files.py

HARD RULE: every rep in REPS must have complete counts (Num for controls;
      Num and Percent_Aged for aged) for BOTH sexes. If ANY are missing, the
      script prints exactly which samples are missing, writes NOTHING, and
      exits non-zero. It will never emit a partial/placeholder design file.
      So add R12's counts to summary_info_v1.xlsx before this will run.
"""

import sys
import openpyxl
from pathlib import Path

XLSX   = Path("helpfiles/AGE_SY/summary_info_v1.xlsx")
OUTDIR = Path("helpfiles/AGE_SY")
REPS   = list(range(1, 13))   # R1–R12

# ── Read Control_Flies ────────────────────────────────────────────────────────
# control[(rep_num, sex)] = Num_Random_Controls
control = {}
ws = openpyxl.load_workbook(XLSX, data_only=True)["Control_Flies"]
first = True
for row in ws.iter_rows(values_only=True):
    if first:
        first = False
        continue
    rep_str, _, sex, _, num = row[:5]
    if not rep_str or not str(rep_str).startswith("Rep"):
        continue
    rep_raw = str(rep_str)[3:]          # "1", "2", ..., "11"
    if not rep_raw.isdigit():           # skip RepA1, RepA2, etc.
        continue
    control[(int(rep_raw), sex)] = num

# ── Read Aged_Flies ───────────────────────────────────────────────────────────
# aged[(rep_num, sex, diet)] = (Num_Aged, Proportion)
aged = {}
ws = openpyxl.load_workbook(XLSX, data_only=True)["Aged_Flies"]
first = True
for row in ws.iter_rows(values_only=True):
    if first:
        first = False
        continue
    rep_str, _, sex, diet, _, _, num_aged, pct_aged = row[:8]
    if not rep_str or not str(rep_str).startswith("Rep"):
        continue
    rep_raw = str(rep_str)[3:]
    if not rep_raw.isdigit():
        continue
    rep_num = int(rep_raw)
    if num_aged in (None, "-") or pct_aged in (None, "-"):
        aged[(rep_num, sex, diet)] = (None, None)
    else:
        aged[(rep_num, sex, diet)] = (int(num_aged), round(float(pct_aged) / 100, 4))

# ── Build rows for one design file ───────────────────────────────────────────
# Appends any samples with missing/incomplete counts to `missing` (they are the
# reason to abort). A row is only produced when its counts are fully present.
def make_rows(diet, sex, missing):
    long_trt = f"AgeSY{diet[2:]}"      # "SY10" → "AgeSY10"
    rows = []

    # Treatment rows first, then control rows
    for long_trt_local, trt_code in [(long_trt, "Z"), ("Con", "C")]:
        for rep in REPS:
            bam = f"{long_trt_local}_R{rep}_{sex}"
            if trt_code == "Z":
                num, prop = aged.get((rep, sex, diet), (None, None))
                if num is None or prop is None:
                    missing.append(f"{bam}  (aged: needs Num_Aged + Percent_Aged)")
                    continue
                num_str, prop_str = str(num), f"{prop:.4f}"
            else:
                num = control.get((rep, sex))
                if num is None:
                    missing.append(f"{bam}  (control: needs Num_Random_Controls)")
                    continue
                num_str, prop_str = str(num), "NA"

            rows.append({
                "bam":     bam,
                "longTRT": long_trt_local,
                "Rep":     f"R{rep}",
                "Sex":     sex,
                "TRT":     trt_code,
                "REP":     rep,
                "Num":     num_str,
                "Prop":    prop_str,
            })
    return rows

# ── Write one design file ─────────────────────────────────────────────────────
HEADER = '"filesize" "bam" "longTRT" "Replicate" "Sex" "TRT" "REP" "REPrep" "Num" "Proportion"'

def write_design(path, rows):
    lines = [HEADER]
    for i, r in enumerate(rows, 1):
        lines.append(
            f'"{i}" 0 "{r["bam"]}" "{r["longTRT"]}" "{r["Rep"]}" '
            f'"{r["Sex"]}" "{r["TRT"]}" {r["REP"]} 1 {r["Num"]} {r["Prop"]}'
        )
    path.write_text("\n".join(lines) + "\n")
    print(f"Wrote {path}  ({len(rows)} rows)")

# ── Build everything in memory; abort HARD if any counts are missing ──────────
# Nothing is written until every rep in REPS has complete counts for every
# diet x sex. One missing value -> list all offenders, write no files, exit 1.
missing = []
outputs = []                                   # (path, rows) pending write
for diet in ["SY10", "SY20"]:
    for sex in ["F", "M"]:
        tag  = diet[2:]                        # "10" or "20"
        rows = make_rows(diet, sex, missing)
        outputs.append((OUTDIR / f"AGE_SY{tag}_{sex}.test.txt", rows))

if missing:
    print(f"\nERROR: incomplete fly-count data in {XLSX}", file=sys.stderr)
    print("Refusing to write design files. Missing counts for:", file=sys.stderr)
    for m in missing:
        print(f"  - {m}", file=sys.stderr)
    print(f"\n{len(missing)} sample(s) missing. Fill them in and rerun. "
          "No files were written.", file=sys.stderr)
    sys.exit(1)

# ── All complete — now (and only now) write all four files ────────────────────
for path, rows in outputs:
    write_design(path, rows)
